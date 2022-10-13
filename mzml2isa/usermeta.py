from __future__ import absolute_import

import json
import openpyxl
import os
import collections

from . import __author__, __license__, __name__, __version__


class UserMetaLoader(object):

    CATEGORIZED_MAP = collections.OrderedDict(
        [
            (
                "Assay Parameters",
                collections.OrderedDict(
                    [
                        (
                            "Chromatography Instrument Name",
                            [("Chromatography Instrument", "name"), False],
                        ),
                        (
                            "Chromatography Instrument Accession Number",
                            [("Chromatography", "accession"), False],
                        ),
                        (
                            "Chromatography Instrument Term Source REF",
                            [("Chromatography Instrument", "ref"), False],
                        ),
                        ("Column model", [("Column model", "value"), False]),
                        ("Column type", [("Column type", "value"), False]),
                        ("Derivatization", [("Derivatization", "value"), False]),
                        ("Post Extraction", [("Post Extraction", "value"), False]),
                    ]
                ),
            ),
            (
                "Characteristics",
                collections.OrderedDict(
                    [
                        (
                            "Organism Name",
                            [("characteristics", "organism", "name"), False],
                        ),
                        (
                            "Organism Accession Number",
                            [("characteristics", "organism", "accession"), False],
                        ),
                        (
                            "Organism Term Source REF",
                            [("characteristics", "organism", "ref"), False],
                        ),
                        (
                            "Organism Part Name",
                            [("characteristics", "organism_part", "name"), False],
                        ),
                        (
                            "Organism Part Accession Number",
                            [("characteristics", "organism_part", "accession"), False],
                        ),
                        (
                            "Organism Part Term Source REF",
                            [("characteristics", "organism_part", "ref"), False],
                        ),
                        (
                            "Organism Variant Name",
                            [("characteristics", "organism_variant", "name"), False],
                        ),
                        (
                            "Organism Variant Accession Number",
                            [
                                ("characteristics", "organism_variant", "accession"),
                                False,
                            ],
                        ),
                        (
                            "Organism Variant Term Source REF",
                            [("characteristics", "organism_variant", "ref"), False],
                        ),
                    ]
                ),
            ),
            (
                "Protocol Description",
                {
                    "Chromatography Description": [("description", "chroma"), False],
                    "Data Transformation Description": [
                        ("description", "data_trans"),
                        False,
                    ],
                    "Extraction Description": [("description", "extraction"), False],
                    "Mass Spectrometry Description": [
                        ("description", "mass_spec"),
                        False,
                    ],
                    "Metabolite Identification Description": [
                        ("description", "metabo_id"),
                        False,
                    ],
                    "Sample Collection Description": [
                        ("description", "sample_collect"),
                        False,
                    ],
                    "Investigation Description": [
                        ("investigation", "description"),
                        False,
                    ],
                },
            ),
            (
                "Investigation",
                collections.OrderedDict(
                    [
                        (
                            "Investigation Identifier",
                            [("investigation", "identifier"), False],
                        ),
                        (
                            "Investigation Release Date",
                            [("investigation", "release_date"), False],
                        ),
                        (
                            "Investigation Submission Date",
                            [("investigation", "submission_date"), False],
                        ),
                        (
                            "Investigation Publication Authors",
                            [("investigation_publication", "author_list"), False],
                        ),
                        (
                            "Investigation Publication Title",
                            [("investigation_publication", "title"), False],
                        ),
                        (
                            "Investigation Publication DOI",
                            [("investigation_publication", "doi"), False],
                        ),
                        (
                            "Investigation Publication Pubmed ID",
                            [("investigation_publication", "pubmed"), False],
                        ),
                        (
                            "Investigation Publication Status Name",
                            [("investigation_publication", "status", "name"), False],
                        ),
                        (
                            "Investigation Publication Status Accession Number",
                            [
                                ("investigation_publication", "status", "accession"),
                                False,
                            ],
                        ),
                        (
                            "Investigation Publication Status Term Source REF",
                            [("investigation_publication", "status", "ref"), False],
                        ),
                    ]
                ),
            ),
            (
                "Study",
                collections.OrderedDict(
                    [
                        ("Study Description", [("study", "description"), False]),
                        ("Study Indentifier", [("study", "identifier"), False]),
                        ("Study Release Date", [("study", "release_date"), False]),
                        (
                            "Study Submission Date",
                            [("study", "submission_date"), False],
                        ),
                        ("Study Title", [("study", "title"), False]),
                        (
                            "Study Publication Authors",
                            [("study_publication", "author_list"), False],
                        ),
                        (
                            "Study Publication Title",
                            [("study_publication", "title"), False],
                        ),
                        (
                            "Study Publication DOI",
                            [("study_publication", "doi"), False],
                        ),
                        (
                            "Study Publication Pubmed ID",
                            [("study_publication", "pubmed"), False],
                        ),
                        (
                            "Study Publication Status Name",
                            [("study_publication", "status", "name"), False],
                        ),
                        (
                            "Study Publication Status Accession Number",
                            [("study_publication", "status", "accession"), False],
                        ),
                        (
                            "Study Publication Status Term Source REF",
                            [("study_publication", "status", "ref"), False],
                        ),
                    ]
                ),
            ),
            (
                "Contacts",
                collections.OrderedDict(
                    [
                        (
                            "Investigation Contacts First Name",
                            [("investigation_contacts", "first_name"), True],
                        ),
                        (
                            "Investigation Contacts Middle Name",
                            [("investigation_contacts", "mid"), True],
                        ),
                        (
                            "Investigation Contacts Last Name",
                            [("investigation_contacts", "last_name"), True],
                        ),
                        (
                            "Investigation Contacts Affiliation",
                            [("investigation_contacts", "affiliation"), True],
                        ),
                        (
                            "Investigation Contacts Adress",
                            [("investigation_contacts", "adress"), True],
                        ),
                        (
                            "Investigation Contacts Email",
                            [("investigation_contacts", "email"), True],
                        ),
                        (
                            "Investigation Contacts Phone",
                            [("investigation_contacts", "phone"), True],
                        ),
                        (
                            "Investigation Contacts Fax",
                            [("investigation_contacts", "fax"), True],
                        ),
                        (
                            "Investigation Contacts Role Name",
                            [("investigation_contacts", "roles", "name"), True],
                        ),
                        (
                            "Investigation Contacts Role Term Source REF",
                            [("investigation_contacts", "roles", "ref"), True],
                        ),
                        (
                            "Investigation Contacts Role Accession Number",
                            [("investigation_contacts", "roles", "accession"), True],
                        ),
                        (
                            "Study Contacts First Name",
                            [("study_contacts", "first_name"), True],
                        ),
                        (
                            "Study Contacts Middle Name",
                            [("study_contacts", "mid"), True],
                        ),
                        (
                            "Study Contacts Last Name",
                            [("study_contacts", "last_name"), True],
                        ),
                        (
                            "Study Contacts Affiliation",
                            [("study_contacts", "affiliation"), True],
                        ),
                        ("Study Contacts Adress", [("study_contacts", "adress"), True]),
                        ("Study Contacts Email", [("study_contacts", "email"), True]),
                        ("Study Contacts Phone", [("study_contacts", "phone"), True]),
                        ("Study Contacts Fax", [("study_contacts", "fax"), True]),
                        (
                            "Study Contacts Role Name",
                            [("study_contacts", "roles", "name"), True],
                        ),
                        (
                            "Study Contacts Role Term Source REF",
                            [("study_contacts", "roles", "ref"), True],
                        ),
                        (
                            "Study Contacts Role Accession Number",
                            [("study_contacts", "roles", "accession"), True],
                        ),
                    ]
                ),
            ),
        ]
    )

    MAP = {
        k: v
        for submap in CATEGORIZED_MAP.values()
        for k, v in submap.items()
    }

    def __init__(self, usermeta_token):
        if usermeta_token is None:
            self.usermeta = None
        elif usermeta_token.endswith(".xlsx"):
            self._parse_xlsx_file(usermeta_token)
        elif usermeta_token.endswith(".json"):
            self._parse_json_file(usermeta_token)
        else:
            self._parse_json_stdin(usermeta_token)

    def _parse_json_file(self, usermeta_token):
        try:
            with open(usermeta_token) as f:
                self.usermeta = json.load(f)
        except json.decoder.JSONDecodeError:
            self.usermeta = None
            warnings.warn(
                "JSON usermeta could not be parsed from {}.".format(usermeta_token)
            )
        except OSError:
            self.usermeta = None
            warnings.warn("File {} not found.".format(usermeta_token))

    def _parse_json_stdin(self, usermeta_token):
        try:
            self.usermeta = json.loads(usermeta_token)
        except json.decoder.JSONDecodeError:
            self.usermeta = None
            warnings.warn("JSON usermeta could not be parsed from <stdin>.")

    def _parse_xlsx_file(self, usermeta_token):
        self.usermeta = {}
        sheet = openpyxl.load_workbook(usermeta_token).worksheets[0]

        for row in sheet.iter_rows():

            # if the row is [Header, Value, None, ..., None] -> non multiple values
            if row[1] != None and all(x.value is None for x in row[2:]):
                header = row[0].value
                value = row[1].value
            # if the row is [Header, Value, None ... None, Value, None ... None, ...] -> multiple values
            else:
                header = row[0].value
                value = [x.value if x.value is not None else "" for x in row[1:]]

            # Skip line if comment or empty headers
            if header is None or header.startswith("#") or not value:
                continue

            # Check in map how to translate excel headers to the metatadata dict
            true_name, more_than_one = self.MAP[header]

            # if there's only one value to write: find the dict to update
            # (self.usermeta[key1][key2][...][keyn])
            if not more_than_one:
                item_to_set = self.usermeta
                for i, path_node in enumerate(true_name[:-1]):
                    item_to_set = item_to_set.setdefault(
                        path_node, {true_name[i + 1]: {}}
                    )
                if isinstance(value, list):
                    item_to_set[true_name[-1]] = ", ".join(value or [])
                else:
                    item_to_set[true_name[-1]] = value or ""

            # if there's only one value to write: find the dict to update
            # (self.usermeta[key1][i][key2][...][keyn] where i is the offset
            # of the current value)
            else:
                for i, value in enumerate(value):
                    item_to_set = self.usermeta.setdefault(true_name[0], [])
                    if len(item_to_set) <= i:
                        item_to_set.append({})
                    item_to_set = self.usermeta[true_name[0]][i]

                    for i, path_node in enumerate(true_name[1:-1]):
                        item_to_set = item_to_set.setdefault(
                            path_node, {true_name[i + 1]: {}}
                        )

                    item_to_set[true_name[-1]] = value

        # Remove empty multiple_values dictionaries
        for mv_key in (v[0][0] for v in self.MAP.values() if v[1]):
            try:
                for value in self.usermeta[mv_key]:
                    empty = not any(
                        v
                        for k, v in value.items()
                        if not isinstance(v, collections.Mapping)
                    )
                    if empty:
                        self.usermeta[mv_key].remove(value)
            except KeyError:
                pass

    @classmethod
    def dump_template_xlsx(cls, output_directory, name="usermeta.xlsx"):

        wb = openpyxl.Workbook()
        x, y = 65, 1

        for category, submap in cls.CATEGORIZED_MAP.items():
            y += 1
            wb.worksheets[0]["{}{}".format(chr(x), y)] = "#### {} ####".format(
                category.upper()
            )
            y += 1
            for header in submap:
                wb.worksheets[0]["{}{}".format(chr(x), y)] = header
                y += 1

        wb.save(os.path.join(output_directory, name))


if __name__ == "__main__":
    UserMetaLoader.dump_template_xlsx(os.getcwd())
